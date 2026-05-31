from __future__ import annotations

import csv
import json
import os
import re
import sys
import webbrowser
from datetime import datetime
from html import escape
from pathlib import Path
from typing import List, Optional
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path("outputs")
MANUAL_REVIEW_ACTIONS = {"manual_reply_if_brand_mention", "pain_point_report", "creator_review"}
COMPETITOR_RESEARCH_WARNING = (
    "Use this dashboard for market research and manual review only. Do not auto DM, "
    "auto follow, auto like, auto reply, or copy competitor content directly."
)
RUN_TYPES = [
    ("listen_", "Social Listening", "social"),
    ("cli_listen_", "Social Listening", "social"),
    ("competitor_", "Competitor Watch", "competitor"),
    ("trends_", "Trend Radar", "trends"),
    ("counts_", "Counts Check", "counts"),
    ("creator_", "Creator Finder", "creator"),
    ("care_", "Customer Care", "care"),
    ("mentions_", "Customer Care", "care"),
]


def now_run_dir(prefix: str = "run") -> Path:
    OUTPUT_DIR.mkdir(exist_ok=True)
    d = OUTPUT_DIR / f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    d.mkdir(parents=True, exist_ok=True)
    return d


def save_csv(path: Path, rows: List[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return
    keys = []
    for r in rows:
        for k in r.keys():
            if k not in keys:
                keys.append(k)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def save_json(path: Path, data) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _int_value(value, default: int = 0) -> int:
    try:
        return int(value or 0)
    except Exception:
        return default


def _dashboard_review_rows(rows: List[dict]) -> List[dict]:
    review_rows = []
    for row in rows:
        action = str(row.get("recommendedAction") or "")
        if action == "no_action_spam":
            continue
        if action in MANUAL_REVIEW_ACTIONS or _int_value(row.get("lead_score")) >= 40:
            review_rows.append(row)
    return review_rows


def _read_csv_rows(path: Path) -> List[dict]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return [dict(row) for row in csv.DictReader(f)]


def _csv_row_count(path: Path) -> int:
    if not path.exists():
        return 0
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return sum(1 for _row in csv.DictReader(f))


def _profiles_by_id(path: Path, profiles: Optional[List[dict]] = None) -> dict:
    rows = profiles or _read_csv_rows(path)
    return {str(row.get("id") or ""): row for row in rows if row.get("id")}


def _post_id(row: dict) -> str:
    return str(row.get("post_id") or row.get("id") or row.get("tweet_id") or "").strip()


def _username(row: dict) -> str:
    return str(row.get("username") or "").strip().lstrip("@")


def _x_post_url(row: dict) -> str:
    post_id = _post_id(row)
    if not post_id:
        return ""
    username = _username(row)
    if username:
        return f"https://x.com/{username}/status/{post_id}"
    return f"https://x.com/i/web/status/{post_id}"


def _x_sample_url(row: dict) -> str:
    sample_url = str(row.get("sample_url") or "").strip()
    if "/status/" in sample_url and ("x.com/" in sample_url or "twitter.com/" in sample_url):
        return sample_url.replace("https://twitter.com/", "https://x.com/")
    post_id = str(row.get("sample_post_id") or row.get("post_id") or row.get("id") or "").strip()
    username = str(row.get("username") or "").strip().lstrip("@")
    if post_id and username:
        return f"https://x.com/{username}/status/{post_id}"
    if post_id:
        return f"https://x.com/i/web/status/{post_id}"
    return ""


def _open_post_link(row: dict, label: str = "Open Post") -> str:
    url = _x_post_url(row)
    if not url:
        return "-"
    return f"<a class='open-link' href='{escape(url)}' target='_blank' rel='noopener noreferrer'>{escape(label)}</a>"


def _competitor_username(row: dict, profiles_by_id: dict) -> str:
    username = _username(row)
    if username:
        return username
    profile = profiles_by_id.get(str(row.get("author_id") or ""))
    return str((profile or {}).get("username") or "").strip().lstrip("@")


def _competitor_post_url(row: dict, username: str) -> str:
    post_id = _post_id(row)
    if username and post_id:
        return f"https://x.com/{username}/status/{post_id}"
    if post_id:
        return f"https://x.com/i/web/status/{post_id}"
    return ""


def _render_competitor_posts_table(rows: List[dict], profiles_by_id: dict) -> str:
    parts = [
        "<div class='card'><h2>Competitor Posts / โพสต์คู่แข่ง</h2>",
        f"<p class='warning'>{escape(COMPETITOR_RESEARCH_WARNING)}</p>",
        "<div class='table-wrap'><table><tr>"
        "<th>username</th><th>created_at</th><th>text</th><th>category</th><th>score</th>"
        "<th>like_count</th><th>reply_count</th><th>retweet_count</th><th>quote_count</th>"
        "<th>bookmark_count</th><th>impression_count</th><th>Open Post</th></tr>",
    ]
    for row in rows:
        username = _competitor_username(row, profiles_by_id)
        url = _competitor_post_url(row, username)
        score = row.get("score", row.get("lead_score", ""))
        open_link = (
            f"<a class='open-link' href='{escape(url)}' target='_blank' rel='noopener noreferrer'>Open Post</a>"
            if url else ""
        )
        parts.append(
            "<tr>"
            f"<td>@{escape(username)}</td>"
            f"<td class='nowrap'>{escape(str(row.get('created_at','')))}</td>"
            f"<td>{escape(str(row.get('text','')))}</td>"
            f"<td>{escape(str(row.get('category','')))}</td>"
            f"<td class='score'>{escape(str(score))}</td>"
            f"<td>{escape(str(row.get('like_count','')))}</td>"
            f"<td>{escape(str(row.get('reply_count','')))}</td>"
            f"<td>{escape(str(row.get('retweet_count','')))}</td>"
            f"<td>{escape(str(row.get('quote_count','')))}</td>"
            f"<td>{escape(str(row.get('bookmark_count','')))}</td>"
            f"<td>{escape(str(row.get('impression_count','')))}</td>"
            f"<td>{open_link}</td>"
            "</tr>"
        )
    parts.append("</table></div></div>")
    return "\n".join(parts)


def _render_social_posts_table(rows: List[dict]) -> str:
    parts = [
        "<div class='card'><h2>Social Listening Posts</h2>",
        "<div class='table-wrap'><table><tr>"
        "<th>score</th><th>category</th><th>username</th><th>created_at</th><th>text</th>"
        "<th>recommendedAction</th><th>like_count</th><th>reply_count</th><th>retweet_count</th>"
        "<th>quote_count</th><th>Open Post</th></tr>",
    ]
    for row in rows:
        parts.append(
            "<tr>"
            f"<td class='score'>{escape(str(row.get('score', row.get('lead_score',''))))}</td>"
            f"<td>{escape(str(row.get('category','')))}</td>"
            f"<td>@{escape(_username(row))}</td>"
            f"<td class='nowrap'>{escape(str(row.get('created_at','')))}</td>"
            f"<td>{escape(str(row.get('text','')))}</td>"
            f"<td>{escape(str(row.get('recommendedAction','')))}</td>"
            f"<td>{escape(str(row.get('like_count','')))}</td>"
            f"<td>{escape(str(row.get('reply_count','')))}</td>"
            f"<td>{escape(str(row.get('retweet_count','')))}</td>"
            f"<td>{escape(str(row.get('quote_count','')))}</td>"
            f"<td>{_open_post_link(row)}</td>"
            "</tr>"
        )
    parts.append("</table></div></div>")
    return "\n".join(parts)


def _render_creator_table(creators: List[dict]) -> str:
    parts = [
        "<div class='card'><h2>Creator Finder</h2>",
        "<div class='table-wrap'><table><tr><th>Score</th><th>User</th><th>Followers</th><th>Posts</th><th>Sample</th><th>Open Sample</th></tr>",
    ]
    for creator in creators[:20]:
        sample_url = _x_sample_url(creator)
        sample_link = (
            f"<a class='open-link' href='{escape(sample_url)}' target='_blank' rel='noopener noreferrer'>Open Sample</a>"
            if sample_url else "-"
        )
        parts.append(
            "<tr>"
            f"<td class='score'>{escape(str(creator.get('creator_score','')))}</td>"
            f"<td>@{escape(str(creator.get('username','')))}</td>"
            f"<td>{escape(str(creator.get('followers_count','')))}</td>"
            f"<td>{escape(str(creator.get('post_count','')))}</td>"
            f"<td>{escape(str(creator.get('sample_text','')))}</td>"
            f"<td>{sample_link}</td>"
            "</tr>"
        )
    parts.append("</table></div></div>")
    return "\n".join(parts)


def _render_lead_queue_table(rows: List[dict]) -> str:
    parts = [
        "<div class='card'><h2>Lead Queue / โพสต์ที่ควรดู</h2>",
        "<div class='table-wrap'><table><tr><th>Score</th><th>หมวด</th><th>User</th><th>ข้อความ</th><th>Recommended Action</th><th>ควรทำอะไร</th><th>ลิงก์</th></tr>",
    ]
    for row in rows:
        parts.append(
            "<tr>"
            f"<td class='score'>{escape(str(row.get('lead_score','')))}</td>"
            f"<td>{escape(str(row.get('category','')))}</td>"
            f"<td>@{escape(_username(row))}</td>"
            f"<td>{escape(str(row.get('text','')))}</td>"
            f"<td>{escape(str(row.get('recommendedAction','')))}</td>"
            f"<td>{escape(str(row.get('action_suggestion','')))}</td>"
            f"<td>{_open_post_link(row)}</td>"
            "</tr>"
        )
    parts.append("</table></div></div>")
    return "\n".join(parts)


def _run_type(folder_name: str) -> tuple[str, str] | None:
    for prefix, label, section in RUN_TYPES:
        if folder_name.startswith(prefix):
            return label, section
    return None


def _run_datetime(run_dir: Path) -> datetime:
    match = re.search(r"(\d{8})_(\d{6})$", run_dir.name)
    if match:
        try:
            return datetime.strptime("_".join(match.groups()), "%Y%m%d_%H%M%S")
        except ValueError:
            pass
    return datetime.fromtimestamp(run_dir.stat().st_mtime)


def _format_run_datetime(run_dir: Path) -> str:
    return _run_datetime(run_dir).strftime("%Y-%m-%d %H:%M:%S")


def _rel_link(path: Path, base_dir: Path) -> str:
    rel = os.path.relpath(path.resolve(), base_dir.resolve()).replace(os.sep, "/")
    encoded = quote(rel, safe="/._-")
    return encoded if encoded.startswith(".") else "./" + encoded


def _load_summary(run_dir: Path) -> dict:
    path = run_dir / "summary.json"
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _read_json(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _session_id_from_path(session_dir: Path) -> str:
    return session_dir.name


def create_research_session(
    outputs_dir: str | Path = OUTPUT_DIR,
    session_name: str = "",
    session_note: str = "",
) -> Path:
    outputs = Path(outputs_dir)
    outputs.mkdir(parents=True, exist_ok=True)
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    session_dir = outputs / f"session_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    session_dir.mkdir(parents=True, exist_ok=True)
    meta = {
        "session_id": _session_id_from_path(session_dir),
        "session_name": (session_name or "").strip() or f"Research Session {created_at}",
        "session_note": (session_note or "").strip(),
        "created_at": created_at,
        "inputs": {},
        "reports": {},
        "report_history": [],
    }
    save_json(session_dir / "session_meta.json", meta)
    return session_dir


def load_session_meta(session_dir: str | Path) -> dict:
    return _read_json(Path(session_dir) / "session_meta.json")


def ensure_research_session(
    outputs_dir: str | Path = OUTPUT_DIR,
    session_name: str = "",
    session_note: str = "",
    current_session_dir: str | Path | None = None,
) -> Path:
    current = Path(current_session_dir) if current_session_dir else None
    requested_name = (session_name or "").strip()
    requested_note = (session_note or "").strip()
    if current and (current / "session_meta.json").exists():
        meta = load_session_meta(current)
        if (
            (meta.get("session_name") or "") == (requested_name or meta.get("session_name") or "")
            and (meta.get("session_note") or "") == requested_note
        ):
            return current
    return create_research_session(outputs_dir, requested_name, requested_note)


def _merge_inputs(existing: dict, updates: dict | None) -> dict:
    merged = dict(existing or {})
    for key, value in (updates or {}).items():
        if value in (None, ""):
            continue
        merged[key] = value
    return merged


def attach_report_to_session(
    session_dir: str | Path,
    report_key: str,
    report_dir: str | Path,
    inputs: Optional[dict] = None,
) -> Path:
    session_path = Path(session_dir)
    report_path = Path(report_dir)
    meta = load_session_meta(session_path)
    if not meta:
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        meta = {
            "session_id": _session_id_from_path(session_path),
            "session_name": f"Research Session {created_at}",
            "session_note": "",
            "created_at": created_at,
            "inputs": {},
            "reports": {},
            "report_history": [],
        }
    reports = dict(meta.get("reports") or {})
    reports[report_key] = report_path.name
    meta["reports"] = reports
    meta["inputs"] = _merge_inputs(meta.get("inputs") or {}, inputs)
    history = list(meta.get("report_history") or [])
    history.append({
        "report_key": report_key,
        "folder": report_path.name,
        "attached_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })
    meta["report_history"] = history
    save_json(session_path / "session_meta.json", meta)
    report_path.mkdir(parents=True, exist_ok=True)
    save_json(report_path / "report_session.json", {
        "session_id": meta.get("session_id", session_path.name),
        "session_name": meta.get("session_name", ""),
        "session_note": meta.get("session_note", ""),
        "created_at": meta.get("created_at", ""),
        "session_folder": session_path.name,
        "report_key": report_key,
    })
    build_session_dashboard(session_path)
    return session_path / "dashboard.html"


def _first_existing_count(run_dir: Path, names: list[str]) -> int:
    for name in names:
        path = run_dir / name
        if path.exists():
            return _csv_row_count(path)
    return 0


def _post_count(run_dir: Path, report_type: str, summary: dict) -> str:
    if summary.get("total_posts") not in (None, ""):
        return str(summary.get("total_posts"))
    if report_type == "Competitor Watch":
        return str(_first_existing_count(run_dir, ["competitor_posts.csv"]))
    if report_type == "Trend Radar":
        return str(_first_existing_count(run_dir, ["trends.csv"]))
    if report_type == "Counts Check":
        return str(_first_existing_count(run_dir, ["counts.csv"]))
    if report_type == "Creator Finder":
        return str(_first_existing_count(run_dir, ["creators.csv"]))
    if report_type == "Customer Care":
        return str(_first_existing_count(run_dir, ["mentions.csv", "care_queue.csv"]))
    return str(_first_existing_count(run_dir, ["posts.csv", "raw_posts_before_filter.csv"]))


def _category_summary(summary: dict) -> str:
    categories = summary.get("categories") or {}
    if not isinstance(categories, dict) or not categories:
        return ""
    return ", ".join(f"{k}: {v}" for k, v in categories.items())


def _report_folder_from_meta(session_dir: Path, meta: dict, key: str) -> Path | None:
    folder = (meta.get("reports") or {}).get(key)
    if not folder:
        return None
    path = session_dir.parent / str(folder)
    return path if path.exists() else None


def _report_card(session_dir: Path, meta: dict, key: str, title: str) -> str:
    run_dir = _report_folder_from_meta(session_dir, meta, key)
    if not run_dir:
        return f"<div class='card'><b>{escape(title)}</b><span>ยังไม่มี</span></div>"
    detected = _run_type(run_dir.name) or (title, key)
    report_type = detected[0]
    summary = _load_summary(run_dir)
    dashboard = run_dir / "dashboard.html"
    link = (
        f"<a class='btn small' href='{escape(_rel_link(dashboard, session_dir))}'>Open Dashboard</a>"
        if dashboard.exists() else ""
    )
    return (
        f"<div class='card'><b>{escape(title)}</b>"
        f"<span>Posts: {escape(_post_count(run_dir, report_type, summary))}</span><br>"
        f"<span>High interest: {escape(str(summary.get('high_interest', '')))}</span><br>"
        f"<small>{escape(run_dir.name)}</small><br>{link}</div>"
    )


def _input_list(value) -> str:
    if isinstance(value, list):
        return "<br>".join(escape(str(item)) for item in value)
    return escape(str(value or ""))


def build_session_dashboard(session_dir: str | Path) -> Path:
    session_path = Path(session_dir)
    meta = load_session_meta(session_path)
    session_path.mkdir(parents=True, exist_ok=True)
    inputs = meta.get("inputs") or {}
    social_dir = _report_folder_from_meta(session_path, meta, "social")
    social_summary = _load_summary(social_dir) if social_dir else {}
    social_posts = _read_csv_rows(social_dir / "posts.csv") if social_dir else []
    social_creators = _read_csv_rows(social_dir / "creators.csv") if social_dir else []
    social_leads = _read_csv_rows(social_dir / "lead_list.csv") if social_dir else []
    competitor_dir = _report_folder_from_meta(session_path, meta, "competitor")
    competitor_table = ""
    if competitor_dir:
        competitor_rows = _read_csv_rows(competitor_dir / "competitor_posts.csv")
        competitor_profiles = _profiles_by_id(competitor_dir / "competitor_profiles.csv")
        if competitor_rows:
            competitor_table = _render_competitor_posts_table(competitor_rows, competitor_profiles)

    report_links = []
    for key, label in [
        ("social", "Social Listening"),
        ("competitor", "Competitor Watch"),
        ("trends", "Trend Radar"),
        ("creator", "Creator Finder"),
        ("care", "Customer Care"),
    ]:
        run_dir = _report_folder_from_meta(session_path, meta, key)
        if not run_dir:
            continue
        dashboard = run_dir / "dashboard.html"
        href = _rel_link(dashboard if dashboard.exists() else run_dir, session_path)
        report_links.append(f"<a class='btn' href='{escape(href)}'>{escape(label)}</a>")

    html = [
        "<!doctype html><html><head><meta charset='utf-8'>",
        "<meta name='viewport' content='width=device-width, initial-scale=1'>",
        f"<title>{escape(str(meta.get('session_name') or 'Research Session'))}</title>",
        "<style>"
        "body{font-family:Segoe UI,Tahoma,sans-serif;margin:0;background:#f6f7fb;color:#111;line-height:1.45}"
        "header{padding:22px 24px 10px}main{padding:0 24px 28px}h1{margin:0 0 6px;font-size:28px}h2{margin:0 0 12px}"
        ".warning{background:#fff3cd;border:1px solid #f1d38a;border-radius:8px;padding:10px 12px;color:#5c4500;margin:12px 0}"
        ".cards{display:grid;grid-template-columns:repeat(3,minmax(190px,1fr));gap:12px;margin:14px 0}.card,section{background:#fff;border-radius:8px;padding:14px;box-shadow:0 1px 8px #ddd}.card b{display:block;margin-bottom:6px}"
        "section{margin:14px 0}.btn{display:inline-block;background:#1f4e78;color:#fff;text-decoration:none;padding:8px 11px;border-radius:6px;margin:3px;white-space:nowrap}.btn.small{font-size:12px;padding:6px 8px}"
        ".table-wrap{overflow-x:auto}table{border-collapse:collapse;width:100%;min-width:1050px}th,td{border:1px solid #ddd;padding:8px;vertical-align:top}th{background:#1f4e78;color:#fff;position:sticky;top:0}.nowrap{white-space:nowrap}.score{font-weight:bold}.open-link{display:inline-block;background:#1f4e78;color:white;text-decoration:none;padding:7px 10px;border-radius:6px;white-space:nowrap}"
        "@media(max-width:900px){header,main{padding-left:14px;padding-right:14px}.cards{grid-template-columns:1fr}h1{font-size:24px}}"
        "</style></head><body><header>",
        f"<h1>{escape(str(meta.get('session_name') or 'Research Session'))}</h1>",
        f"<p><b>Session ID:</b> {escape(str(meta.get('session_id') or session_path.name))}</p>",
        f"<p><b>Created at:</b> {escape(str(meta.get('created_at') or ''))}</p>",
        f"<p><b>Note:</b> {escape(str(meta.get('session_note') or ''))}</p>",
        f"<p class='warning'>{escape(COMPETITOR_RESEARCH_WARNING)}</p>",
        "<div class='cards'>",
        _report_card(session_path, meta, "social", "Social Listening"),
        _report_card(session_path, meta, "competitor", "Competitor Watch"),
        _report_card(session_path, meta, "trends", "Trend Radar"),
        "</div></header><main>",
        "<section><h2>Inputs / ข้อมูลที่ค้นหา</h2>",
        f"<p><b>Social query:</b><br>{_input_list(inputs.get('social_queries'))}</p>",
        f"<p><b>Competitor usernames:</b><br>{_input_list(inputs.get('competitor_usernames'))}</p>",
        f"<p><b>Trend WOEID:</b> {escape(str(inputs.get('trend_woeid') or ''))}</p>",
        f"<p>{''.join(report_links) if report_links else 'ยังไม่มีรายงานใน session นี้'}</p>",
        "</section>",
    ]
    if social_dir:
        html.extend([
            "<section><h2>Social Summary</h2>",
            "<div class='cards'>",
            f"<div class='card'><b>โพสต์ทั้งหมด</b><span style='font-size:28px'>{escape(str(social_summary.get('total_posts', len(social_posts))))}</span></div>",
            f"<div class='card'><b>ความสนใจสูง</b><span style='font-size:28px'>{escape(str(social_summary.get('high_interest', 0)))}</span></div>",
            f"<div class='card'><b>หมวดโพสต์</b>{escape(_category_summary(social_summary) or '-')}</div>",
            "</div></section>",
        ])
        if social_posts:
            html.append(_render_social_posts_table(social_posts))
        if social_creators:
            html.append(_render_creator_table(social_creators))
        html.append(_render_lead_queue_table(social_leads or _dashboard_review_rows(social_posts)))
    if competitor_table:
        html.append(competitor_table)
    html.append("</main></body></html>")
    path = session_path / "dashboard.html"
    path.write_text("\n".join(html), encoding="utf-8")
    return path


def _link_buttons(paths: List[Path], base_dir: Path, label: str) -> str:
    buttons = []
    for path in paths:
        buttons.append(
            f"<a class='btn small' href='{escape(_rel_link(path, base_dir))}'>{escape(label if len(paths) == 1 else path.name)}</a>"
        )
    return " ".join(buttons)


def _run_record(run_dir: Path, outputs_dir: Path) -> dict | None:
    detected = _run_type(run_dir.name)
    if not detected:
        return None
    report_type, section = detected
    summary = _load_summary(run_dir)
    session = _read_json(run_dir / "report_session.json")
    csv_files = sorted(run_dir.glob("*.csv"))
    xlsx_files = sorted(run_dir.glob("*.xlsx"))
    dashboard = run_dir / "dashboard.html"
    return {
        "date_text": _format_run_datetime(run_dir),
        "timestamp": _run_datetime(run_dir).timestamp(),
        "report_type": report_type,
        "section": section,
        "folder": run_dir.name,
        "session_id": str(session.get("session_id") or ""),
        "session_name": str(session.get("session_name") or ""),
        "session_note": str(session.get("session_note") or ""),
        "post_count": _post_count(run_dir, report_type, summary),
        "high_interest": str(summary.get("high_interest", "")),
        "categories": _category_summary(summary),
        "dashboard_link": (
            f"<a class='btn' href='{escape(_rel_link(dashboard, outputs_dir))}'>Open Dashboard</a>"
            if dashboard.exists() else ""
        ),
        "csv_links": _link_buttons(csv_files, outputs_dir, "Open CSV") if csv_files else "",
        "xlsx_links": _link_buttons(xlsx_files, outputs_dir, "Open XLSX") if xlsx_files else "",
        "folder_link": f"<a class='btn muted' href='{escape(_rel_link(run_dir, outputs_dir))}/'>Open Folder</a>",
    }


def _render_hub_rows(records: List[dict]) -> str:
    if not records:
        return "<p class='empty'>ยังไม่มีรายงานในหมวดนี้</p>"
    parts = [
        "<div class='table-wrap'><table><tr>"
        "<th>Date/time</th><th>Report type</th><th>Session name</th><th>Note</th><th>Folder</th><th>Post count</th>"
        "<th>High interest</th><th>Category summary</th><th>Dashboard</th><th>CSV</th><th>XLSX</th><th>Folder</th></tr>"
    ]
    for record in records:
        parts.append(
            "<tr>"
            f"<td class='nowrap'>{escape(record['date_text'])}</td>"
            f"<td>{escape(record['report_type'])}</td>"
            f"<td>{escape(record.get('session_name',''))}</td>"
            f"<td>{escape(record.get('session_note',''))}</td>"
            f"<td><code>{escape(record['folder'])}</code></td>"
            f"<td>{escape(record['post_count'])}</td>"
            f"<td>{escape(record['high_interest'])}</td>"
            f"<td>{escape(record['categories'])}</td>"
            f"<td>{record['dashboard_link']}</td>"
            f"<td>{record['csv_links']}</td>"
            f"<td>{record['xlsx_links']}</td>"
            f"<td>{record['folder_link']}</td>"
            "</tr>"
        )
    parts.append("</table></div>")
    return "\n".join(parts)


def _session_report_link(meta: dict, outputs_dir: Path, key: str, label: str) -> str:
    folder = (meta.get("reports") or {}).get(key)
    if not folder:
        return ""
    path = outputs_dir / str(folder) / "dashboard.html"
    if not path.exists():
        path = outputs_dir / str(folder)
    return f"<a class='btn small' href='{escape(_rel_link(path, outputs_dir))}'>{escape(label)}</a>"


def _session_post_summary(meta: dict, outputs_dir: Path) -> str:
    parts = []
    for key, label in [("social", "Social"), ("competitor", "Competitor"), ("trends", "Trend"), ("creator", "Creator"), ("care", "Care")]:
        folder = (meta.get("reports") or {}).get(key)
        if not folder:
            continue
        run_dir = outputs_dir / str(folder)
        detected = _run_type(run_dir.name)
        report_type = detected[0] if detected else label
        parts.append(f"{label}: {_post_count(run_dir, report_type, _load_summary(run_dir))}")
    return ", ".join(parts)


def _session_record(session_dir: Path, outputs_dir: Path) -> dict | None:
    meta = load_session_meta(session_dir)
    if not meta:
        return None
    dashboard = session_dir / "dashboard.html"
    if not dashboard.exists():
        build_session_dashboard(session_dir)
    reports = meta.get("reports") or {}
    included = ", ".join(k for k in ["social", "competitor", "trends", "creator", "care"] if reports.get(k))
    return {
        "date_text": str(meta.get("created_at") or _format_run_datetime(session_dir)),
        "timestamp": _run_datetime(session_dir).timestamp(),
        "session_name": str(meta.get("session_name") or ""),
        "session_note": str(meta.get("session_note") or ""),
        "included_reports": included,
        "post_summary": _session_post_summary(meta, outputs_dir),
        "dashboard_link": f"<a class='btn' href='{escape(_rel_link(dashboard, outputs_dir))}'>Open Session Dashboard</a>",
        "social_link": _session_report_link(meta, outputs_dir, "social", "Open Social"),
        "competitor_link": _session_report_link(meta, outputs_dir, "competitor", "Open Competitor"),
        "trend_link": _session_report_link(meta, outputs_dir, "trends", "Open Trend"),
        "folder_link": f"<a class='btn muted' href='{escape(_rel_link(session_dir, outputs_dir))}/'>Open Folder</a>",
    }


def _render_session_rows(records: List[dict]) -> str:
    if not records:
        return "<p class='empty'>ยังไม่มี Research Session</p>"
    parts = [
        "<div class='table-wrap'><table><tr>"
        "<th>Date/time</th><th>Session name</th><th>Note</th><th>Included reports</th>"
        "<th>Post count summary</th><th>Open Session Dashboard</th><th>Open Social</th>"
        "<th>Open Competitor</th><th>Open Trend</th><th>Open Folder</th></tr>"
    ]
    for record in records:
        parts.append(
            "<tr>"
            f"<td class='nowrap'>{escape(record['date_text'])}</td>"
            f"<td>{escape(record['session_name'])}</td>"
            f"<td>{escape(record['session_note'])}</td>"
            f"<td>{escape(record['included_reports'])}</td>"
            f"<td>{escape(record['post_summary'])}</td>"
            f"<td>{record['dashboard_link']}</td>"
            f"<td>{record['social_link']}</td>"
            f"<td>{record['competitor_link']}</td>"
            f"<td>{record['trend_link']}</td>"
            f"<td>{record['folder_link']}</td>"
            "</tr>"
        )
    parts.append("</table></div>")
    return "\n".join(parts)


def build_dashboard_hub(outputs_dir: str | Path = OUTPUT_DIR) -> Path:
    outputs = Path(outputs_dir)
    outputs.mkdir(parents=True, exist_ok=True)
    records = []
    session_records = []
    for run_dir in outputs.iterdir():
        if not run_dir.is_dir():
            continue
        if run_dir.name.startswith("session_"):
            session_record = _session_record(run_dir, outputs)
            if session_record:
                session_records.append(session_record)
        else:
            record = _run_record(run_dir, outputs)
            if record:
                records.append(record)
    records.sort(key=lambda item: item["timestamp"], reverse=True)
    session_records.sort(key=lambda item: item["timestamp"], reverse=True)

    latest_social = next((r for r in records if r["section"] == "social"), None)
    latest_competitor = next((r for r in records if r["section"] == "competitor"), None)
    latest_trends = next((r for r in records if r["section"] == "trends"), None)
    latest_any = records[0] if records else None

    sections = [
        ("sessions", "Research Sessions", session_records),
        ("all", "All Reports", records),
        ("social", "Social Listening", [r for r in records if r["section"] == "social"]),
        ("competitor", "Competitor Watch", [r for r in records if r["section"] == "competitor"]),
        ("trends", "Trend Radar", [r for r in records if r["section"] == "trends"]),
        ("creator", "Creator Finder", [r for r in records if r["section"] == "creator"]),
        ("care", "Customer Care", [r for r in records if r["section"] == "care"]),
    ]

    def latest_text(record: dict | None) -> str:
        if not record:
            return "ยังไม่มี"
        return f"{record['date_text']}<br><small>{escape(record['folder'])}</small>"

    html = [
        "<!doctype html><html><head><meta charset='utf-8'>",
        "<meta name='viewport' content='width=device-width, initial-scale=1'>",
        "<title>BN9 Dashboard Hub</title>",
        "<style>"
        "body{font-family:Segoe UI,Tahoma,sans-serif;margin:0;background:#f6f7fb;color:#111;line-height:1.45}"
        "header{padding:22px 24px 12px}main{padding:0 24px 28px}h1{margin:0 0 6px;font-size:28px}h2{margin:0 0 12px}"
        ".warning{background:#fff3cd;border:1px solid #f1d38a;border-radius:8px;padding:10px 12px;color:#5c4500;margin:12px 0}"
        ".cards{display:grid;grid-template-columns:repeat(5,minmax(170px,1fr));gap:12px;margin:14px 0}"
        ".card{background:#fff;border-radius:8px;padding:14px;box-shadow:0 1px 8px #ddd}.card b{display:block;margin-bottom:6px}"
        ".nav{display:flex;gap:8px;flex-wrap:wrap;margin:16px 0}.nav a,.btn{display:inline-block;background:#1f4e78;color:#fff;text-decoration:none;padding:8px 11px;border-radius:6px;white-space:nowrap}"
        ".btn.small{font-size:12px;padding:6px 8px;margin:2px}.btn.muted{background:#59636e}section{background:#fff;border-radius:8px;padding:16px;margin:14px 0;box-shadow:0 1px 8px #ddd}"
        ".table-wrap{overflow-x:auto;max-height:70vh}table{border-collapse:collapse;width:100%;min-width:1200px}th,td{border:1px solid #ddd;padding:8px;vertical-align:top}th{background:#1f4e78;color:#fff;position:sticky;top:0}.nowrap{white-space:nowrap}code{font-family:Consolas,monospace}.empty{color:#666}"
        "@media(max-width:900px){header,main{padding-left:14px;padding-right:14px}.cards{grid-template-columns:1fr 1fr}h1{font-size:24px}}"
        "@media(max-width:560px){.cards{grid-template-columns:1fr}.nav a,.btn{width:auto}.card{padding:12px}body{font-size:15px}}"
        "</style>",
        "</head><body><header>",
        "<h1>BN9 Dashboard Hub</h1>",
        f"<p>สร้างเมื่อ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>",
        f"<p class='warning'>{escape(COMPETITOR_RESEARCH_WARNING)}</p>",
        "<div class='cards'>",
        f"<div class='card'><b>Latest Social Listening</b>{latest_text(latest_social)}</div>",
        f"<div class='card'><b>Latest Competitor Watch</b>{latest_text(latest_competitor)}</div>",
        f"<div class='card'><b>Latest Trend Radar</b>{latest_text(latest_trends)}</div>",
        f"<div class='card'><b>Total report folders</b><span style='font-size:28px'>{len(records)}</span></div>",
        f"<div class='card'><b>Latest generated time</b>{latest_text(latest_any)}</div>",
        "</div>",
        "<nav class='nav'>",
    ]
    for section_id, label, _records in sections:
        html.append(f"<a href='#{section_id}'>{escape(label)}</a>")
    html.extend(["</nav></header><main>"])
    for section_id, label, section_records in sections:
        html.append(f"<section id='{escape(section_id)}'><h2>{escape(label)}</h2>")
        if section_id == "sessions":
            html.append(_render_session_rows(section_records))
        else:
            html.append(_render_hub_rows(section_records))
        html.append("</section>")
    html.append("</main></body></html>")

    hub_path = outputs / "dashboard_hub.html"
    hub_path.write_text("\n".join(html), encoding="utf-8")
    return hub_path


def _sheet_from_rows(wb: Workbook, title: str, rows: List[dict]) -> None:
    ws = wb.create_sheet(title=title[:31])
    if not rows:
        ws.append(["ไม่มีข้อมูล"])
        return
    keys = []
    for r in rows:
        for k in r.keys():
            if k not in keys:
                keys.append(k)
    ws.append(keys)
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for r in rows:
        values = []
        for k in keys:
            value = r.get(k, "")
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            values.append(value)
        ws.append(values)
    ws.freeze_panes = "A2"
    for idx, key in enumerate(keys, start=1):
        width = min(60, max(12, len(key) + 2))
        if key in {"text", "reply_draft", "action_suggestion", "recommendedAction", "sample_text", "user_description"}:
            width = 45
        if key == "url":
            width = 42
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def save_excel(path: Path, sheets: dict[str, List[dict]], summary: Optional[dict] = None) -> None:
    wb = Workbook()
    default = wb.active
    default.title = "Summary"
    default.append(["หัวข้อ", "ค่า"])
    default["A1"].font = Font(bold=True, color="FFFFFF")
    default["B1"].font = Font(bold=True, color="FFFFFF")
    default["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    default["B1"].fill = PatternFill("solid", fgColor="1F4E78")
    if summary:
        for k, v in summary.items():
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)
            default.append([k, v])
    default.column_dimensions["A"].width = 28
    default.column_dimensions["B"].width = 80
    for title, rows in sheets.items():
        _sheet_from_rows(wb, title, rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def save_dashboard(
    path: Path,
    title: str,
    summary: dict,
    rows: List[dict],
    creators: Optional[List[dict]] = None,
    trends: Optional[List[dict]] = None,
    competitor_profiles: Optional[List[dict]] = None,
) -> None:
    cats = summary.get("categories", {}) or {}
    ideas = summary.get("content_ideas", []) or []
    top = _dashboard_review_rows(rows)[:20]
    creators = creators or []
    trends = trends or []
    is_competitor_dashboard = "competitor" in title.lower()
    is_social_dashboard = "social listening" in title.lower() or "csv analysis" in title.lower()
    competitor_profiles_by_id = _profiles_by_id(path.parent / "competitor_profiles.csv", competitor_profiles) if is_competitor_dashboard else {}
    html = [
        "<!doctype html><html><head><meta charset='utf-8'>",
        f"<title>{escape(title)}</title>",
        "<meta name='viewport' content='width=device-width, initial-scale=1'>",
        "<style>body{font-family:Segoe UI,Tahoma,sans-serif;margin:24px;background:#f6f7fb;color:#111;line-height:1.45}.card{background:white;border-radius:8px;padding:18px;margin:12px 0;box-shadow:0 1px 8px #ddd;overflow-x:auto}table{border-collapse:collapse;width:100%;background:white}th,td{border:1px solid #ddd;padding:8px;vertical-align:top}th{background:#1f4e78;color:white}.kpi{display:inline-block;background:#fff;padding:18px;margin:8px;border-radius:8px;box-shadow:0 1px 8px #ddd;min-width:150px}.score{font-weight:bold}.warning{background:#fff3cd;border:1px solid #f1d38a;border-radius:8px;padding:10px 12px;color:#5c4500}.table-wrap{overflow-x:auto}.nowrap{white-space:nowrap}.open-link{display:inline-block;background:#1f4e78;color:white;text-decoration:none;padding:7px 10px;border-radius:6px;white-space:nowrap}@media(max-width:760px){body{margin:14px;font-size:15px}.kpi{display:block;margin:8px 0}.card{padding:14px}table{min-width:900px}}</style>",
        "</head><body>",
        f"<h1>{escape(title)}</h1>",
        f"<p>สร้างเมื่อ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>",
        "<div>",
        f"<div class='kpi'><b>โพสต์ทั้งหมด</b><br><span style='font-size:28px'>{summary.get('total_posts', len(rows))}</span></div>",
        f"<div class='kpi'><b>ความสนใจสูง</b><br><span style='font-size:28px'>{summary.get('high_interest', 0)}</span></div>",
        f"<div class='kpi'><b>ความสนใจกลาง</b><br><span style='font-size:28px'>{summary.get('medium_interest', 0)}</span></div>",
        "</div>",
        "<div class='card'><h2>หมวดโพสต์</h2><ul>",
    ]
    for k, v in cats.items():
        html.append(f"<li><b>{escape(str(k))}</b>: {escape(str(v))}</li>")
    html.append("</ul></div>")
    html.append("<div class='card'><h2>ไอเดียคอนเทนต์</h2><ol>")
    for idea in ideas:
        html.append(f"<li>{escape(str(idea))}</li>")
    html.append("</ol></div>")
    if trends:
        html.append("<div class='card'><h2>Trend Radar</h2><table><tr><th>Trend</th><th>Tweet count</th></tr>")
        for r in trends[:30]:
            html.append(f"<tr><td>{escape(str(r.get('trend_name','')))}</td><td>{escape(str(r.get('tweet_count','')))}</td></tr>")
        html.append("</table></div>")
    if is_social_dashboard and rows:
        html.append(_render_social_posts_table(rows))
    if creators:
        html.append(_render_creator_table(creators))
    if is_competitor_dashboard:
        html.append(_render_competitor_posts_table(rows, competitor_profiles_by_id))
    html.append(_render_lead_queue_table(top))
    html.append("</body></html>")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(html), encoding="utf-8")


def open_path(path: Path) -> None:
    path = path.resolve()
    if sys.platform.startswith("win"):
        os.startfile(str(path))  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        os.system(f"open {path!s}")
    else:
        os.system(f"xdg-open {path!s}")


def open_browser(path: Path) -> None:
    webbrowser.open(path.resolve().as_uri())
